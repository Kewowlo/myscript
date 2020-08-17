# -*- coding: utf-8 -*- 
import os
import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException

def open_xls_as_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.get_active_sheet()

    for row in xrange(0, nrows):
        for col in xrange(0, ncols):
            sheet1.cell(row=row+1, column=col+1).value = sheet.cell_value(row, col)

    return book1
if __name__ == "__main__":
    
    fpath = r"C:\Users\kewowlo\Desktop\0729\二维码门牌".decode('utf-8')

    for root, dirs, files in os.walk(fpath):
        for filename in files:
            fileroot = os.path.join(root,filename)
            print fileroot
            newbook = open_xls_as_xlsx(fileroot)
            newbook.save(fileroot+"x")
    
    print "--end--------"